import openpyxl

#file-workbook-sheets-rows-cells
file = "F:\selenium youtube\Trail.xlsx"
workbook = openpyxl.load_workbook(file)
#sheet=workbbok.active : gets active sheet from excel if file has onlu one sheet
sheet = workbook["Sheet1"]
 
rows = sheet.max_row #counts no of rows in a excel sheet
column = sheet.max_column

#read all the rows and columns data
for r in range (1,rows+1):
    for c in range (1,column+1):
        print(sheet.cell(r,c).value,end= '        ')
    print()
