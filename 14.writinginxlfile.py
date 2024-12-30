import openpyxl

#writing same data in xl file
file = "F:\selenium youtube\Book1.xlsx"

workbook = openpyxl.load_workbook(file)
sheet = workbook.active
# for r in range(1,6):
#     for c in range (1,4):
#         sheet.cell(r,c).value="Milan"
# workbook.save(file)

#writing in each col and row individually

sheet.cell(1,1).value=123
sheet.cell(1,2).value="Melon"
sheet.cell(1,3).value="Financial Advisor"

sheet.cell(2,1).value=456
sheet.cell(2,2).value="Milan"
sheet.cell(2,3).value="engineer"

sheet.cell(3,1).value=589
sheet.cell(3,2).value="Me-elon"
sheet.cell(3,3).value="QA"

workbook.save(file)